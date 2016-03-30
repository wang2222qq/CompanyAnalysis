#!/usr/bin/python3
#-*- coding:utf-8 -*-

__author__ = 'Francics'

import os.path
try:
   from openpyxl import Workbook,worksheet,load_workbook
   from collections import Iterable
except ImportError as e:
   print("电脑上没有安装模块：%s" % e )

class WriteFile():
    def __init__(self):
        self.__workbook = None
        self.__worksheet = None
        self.__filename = None
        self.__row = 0

    def Open(self,filename):
        if os.path.exists(filename):
            self.__workbook = load_workbook(filename)
        else:
            self.__workbook = Workbook()
        self.__filename = filename

    def Close(self):
        self.__workbook = None

    def setSheet(self,sheetname):
        if not self.__workbook:
            raise ValueError("you need open a workbook")
        try:
            self.__worksheet = self.__workbook[sheetname]
        except KeyError:
            ##在文件最后追加工作簿
            self.__worksheet = self.__workbook.create_sheet()
            self.__worksheet.title = sheetname

    def WriteLine(self,mrow,line):
        if not self.__worksheet:
            raise ValueError("you need open select a worksheet")

        if not isinstance(line,Iterable):
            raise TypeError("line must be Iterable!")

        if not isinstance(mrow,int):
            raise TypeError("row must be int")

        if mrow == 1 and len(self.__worksheet.rows[0]) > 0:
            raise ValueError("sheet:[%s] first line have value!")

        icolumn=1
        for word in line:
            _ = self.__worksheet.cell(row=mrow,column=icolumn,value=word)
            icolumn += 1

    def Save(self):
        if not self.__workbook:
            raise ValueError("you need open a workbook")
        self.__workbook.save(self.__filename)

    @property
    def rows(self):
        if self.__worksheet:
            return self.__worksheet.rows
        else:
            return 0

    @property
    def cols(self):
        if self.__worksheet:
            return self.__worksheet.columns
        else:
            return 0